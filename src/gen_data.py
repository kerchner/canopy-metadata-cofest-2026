#!/usr/bin/env python3
"""Generate the synthetic example study (SPbE-2026) for the Canopy AI-metadata CoFest.

Outputs (into ../data/synthetic-study/):
  - SPbE-2026_dataset.xlsx   measurements + data_dictionary sheets
  - SPbE-2026_dataset.csv    same measurements, CSV (no licensing concerns)
  - SPbE-2026_protocol.pdf   ~10-page study protocol
  - SPbE-2026_SOP_sample-collection.pdf

The 20 dataset columns deliberately mix CEDAR field types — numeric, date,
boolean/checkbox, ontology-controlled (with branches: country, disease), and
external-authority identifiers (ORCID, PubMed ID, DOI) — so the example
exercises interesting metadata, not just strings. All data is fictional.
"""
import csv
import random
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, ListFlowable, ListItem, PageBreak)

random.seed(42)
GREEN = "1F6F54"
OUT = Path("/sessions/sharp-intelligent-mendel/mnt/CANOPY_PROD/canopy-metadata-cofest-2026/data/synthetic-study")
OUT.mkdir(parents=True, exist_ok=True)

STUDY_ID = "SPbE-2026"
STUDY_TITLE = ("Effect of 12-Week Time-Restricted Eating on Glycemic Control "
               "in Prediabetic Adults: A Synthetic Pilot Study")
PROTOCOL_DOI = "10.5281/zenodo.0000000"      # fictional
REFERENCE_PMID = "31813466"                  # a real TRE review PMID, used illustratively
ORCIDS = ["0000-0002-1825-0097", "0000-0001-5109-3700", "0000-0002-9876-5432"]
COUNTRIES = ["United States", "Germany", "Spain", "Canada"]  # GAZ / ISO-3166 controlled

# ---------------------------------------------------------------- schema (20 cols)
# (name, type, controlled/authority, description)
SCHEMA = [
    ("subject_id",            "string",   "",                       "Subject identifier"),
    ("enrollment_date",       "date",     "",                       "Date of enrollment (ISO-8601)"),
    ("age_years",             "numeric",  "",                       "Age at enrollment"),
    ("sex",                   "controlled","NCIT (sex)",            "Biological sex"),
    ("country",               "controlled","GAZ / ISO-3166 branch", "Country of the enrolling site"),
    ("condition",             "controlled","MONDO branch",          "Condition under study"),
    ("study_arm",             "controlled","NCIT (treatment arm)",  "Assigned study arm"),
    ("bmi_kg_m2",             "numeric",  "",                       "Body mass index"),
    ("baseline_glucose_mg_dl","numeric",  "",                       "Fasting plasma glucose, baseline"),
    ("week12_glucose_mg_dl",  "numeric",  "",                       "Fasting plasma glucose, week 12"),
    ("hba1c_pct",             "numeric",  "",                       "HbA1c at week 12"),
    ("systolic_bp_mmhg",      "numeric",  "",                       "Systolic blood pressure, week 12"),
    ("on_glucose_medication", "boolean",  "",                       "On glucose-lowering medication?"),
    ("adverse_event",         "boolean",  "",                       "Any adverse event reported?"),
    ("completed_study",       "boolean",  "",                       "Completed through week 12?"),
    ("visit_date",            "date",     "",                       "Final visit date (ISO-8601)"),
    ("adherence_pct",         "numeric",  "",                       "Protocol adherence"),
    ("investigator_orcid",    "identifier","ORCID",                 "ORCID iD of enrolling investigator"),
    ("reference_pmid",        "identifier","PubMed ID",             "PubMed ID of the reference publication"),
    ("protocol_doi",          "identifier","DOI",                   "DOI of the study protocol"),
]
HEADERS = [c[0] for c in SCHEMA]

def make_row(i):
    arm = "TRE" if i % 2 == 0 else "control"
    bg = random.randint(102, 124)
    drop = random.randint(6, 16) if arm == "TRE" else random.randint(-2, 6)
    wg = max(88, bg - drop)
    completed = random.random() > 0.07
    enroll_m = random.randint(1, 2)
    return [
        f"SUBJ-{i:03d}",
        f"2026-0{enroll_m}-{random.randint(5,27):02d}",
        random.randint(40, 65),
        random.choice(["female", "male"]),
        random.choice(COUNTRIES),
        "prediabetes",
        arm,
        round(random.uniform(26.0, 34.0), 1),
        bg,
        (wg if completed else ""),
        (round(random.uniform(5.3, 6.4), 1) if completed else ""),
        random.randint(112, 138),
        random.choice([False, False, False, True]),
        random.choice([False, False, False, True]),
        completed,
        (f"2026-0{enroll_m+3}-{random.randint(5,27):02d}" if completed else ""),
        (random.randint(70, 99) if completed else ""),
        random.choice(ORCIDS),
        REFERENCE_PMID,
        PROTOCOL_DOI,
    ]

ROWS = [make_row(i) for i in range(1, 41)]   # 40 subjects

# ---------------------------------------------------------------- CSV
csv_path = OUT / "SPbE-2026_dataset.csv"
with open(csv_path, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(HEADERS)
    w.writerows(ROWS)
print("wrote", csv_path)

# ---------------------------------------------------------------- XLSX
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "measurements"
hdr_fill = PatternFill("solid", fgColor=GREEN)
hdr_font = Font(bold=True, color="FFFFFF")
for c, name in enumerate(HEADERS, 1):
    cell = ws.cell(1, c, name)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
for r, row in enumerate(ROWS, 2):
    for c, val in enumerate(row, 1):
        ws.cell(r, c, val)
ws.freeze_panes = "B2"
for col in ws.columns:
    width = max(len(str(x.value)) for x in col if x.value is not None) + 2
    ws.column_dimensions[col[0].column_letter].width = min(width, 24)

# data dictionary sheet
dd = wb.create_sheet("data_dictionary")
dd_headers = ["column_name", "field_type", "controlled_by / authority", "units", "example", "ontology_reference"]
ONT = {
    "sex": "NCIT:C28421 (Sex); values female=NCIT:C16576, male=NCIT:C20197",
    "country": "GAZ / ISO-3166-1 (e.g. United States = GAZ:00002459)",
    "condition": "MONDO:0005827 (prediabetes)",
    "study_arm": "NCIT:C174266 (Study Arm)",
    "bmi_kg_m2": "NCIT:C16358 (Body Mass Index); UO:0000077 (kg/m^2)",
    "baseline_glucose_mg_dl": "NCIT:C105585 (Fasting Blood Glucose); mg/dL",
    "week12_glucose_mg_dl": "NCIT:C105585 (Fasting Blood Glucose); mg/dL",
    "hba1c_pct": "NCIT:C64849 (Hemoglobin A1C); %",
    "systolic_bp_mmhg": "NCIT:C25298 (Systolic Blood Pressure); mmHg",
    "adverse_event": "NCIT:C41331 (Adverse Event)",
    "investigator_orcid": "ORCID identifier (https://orcid.org/)",
    "reference_pmid": "PubMed identifier (https://pubmed.ncbi.nlm.nih.gov/)",
    "protocol_doi": "DOI (https://doi.org/)",
    "enrollment_date": "ISO-8601 date",
    "visit_date": "ISO-8601 date",
    "age_years": "years",
    "adherence_pct": "% (0-100)",
}
UNITS = {"bmi_kg_m2": "kg/m^2", "baseline_glucose_mg_dl": "mg/dL", "week12_glucose_mg_dl": "mg/dL",
         "hba1c_pct": "%", "systolic_bp_mmhg": "mmHg", "age_years": "years", "adherence_pct": "%"}
EX = {r[0]: r for r in [ROWS[0]]}  # first row for examples
example_row = ROWS[0]
for c, h in enumerate(dd_headers, 1):
    cell = dd.cell(1, c, h); cell.fill = hdr_fill; cell.font = hdr_font
for r, (name, ftype, auth, desc) in enumerate(SCHEMA, 2):
    ex = example_row[HEADERS.index(name)]
    dd.cell(r, 1, name)
    dd.cell(r, 2, ftype)
    dd.cell(r, 3, auth)
    dd.cell(r, 4, UNITS.get(name, ""))
    dd.cell(r, 5, str(ex))
    dd.cell(r, 6, ONT.get(name, ""))
for col in dd.columns:
    width = max(len(str(x.value)) for x in col if x.value is not None) + 2
    dd.column_dimensions[col[0].column_letter].width = min(width, 55)

xlsx_path = OUT / "SPbE-2026_dataset.xlsx"
wb.save(xlsx_path)
print("wrote", xlsx_path)

# ---------------------------------------------------------------- PDF helpers
styles = getSampleStyleSheet()
styles.add(ParagraphStyle("H", parent=styles["Heading1"], textColor=colors.HexColor("#"+GREEN), spaceBefore=6))
styles.add(ParagraphStyle("H2b", parent=styles["Heading2"], textColor=colors.HexColor("#"+GREEN), spaceBefore=10))
styles.add(ParagraphStyle("H3b", parent=styles["Heading3"], textColor=colors.HexColor("#2a2a2a")))
body = styles["BodyText"]; body.spaceAfter = 8; body.leading = 14

def build_pdf(path, flow):
    SimpleDocTemplate(str(path), pagesize=LETTER, title=STUDY_TITLE,
                      topMargin=0.9*inch, bottomMargin=0.9*inch,
                      leftMargin=1*inch, rightMargin=1*inch).build(flow)
    print("wrote", path)

def P(t, s="BodyText"): return Paragraph(t, styles[s])
def bullets(items, kind="bullet"):
    return ListFlowable([ListItem(P(i)) for i in items], bulletType=kind, leftIndent=18)

# ---------------------------------------------------------------- protocol (~10 pages)
f = []
f += [
    Spacer(1, 60),
    P("STUDY PROTOCOL", "H"),
    P(f"<b>{STUDY_TITLE}</b>", "Heading2"),
    Spacer(1, 18),
    P(f"Protocol identifier: <b>{STUDY_ID}</b>"),
    P("Protocol version: 1.0"),
    P(f"Protocol DOI: {PROTOCOL_DOI}"),
    P(f"Reference publication (PubMed ID): {REFERENCE_PMID}"),
    P("Sponsor: Synthetic Data Working Group, Canopy Project"),
    P("Sites: multi-site (United States, Germany, Spain, Canada)"),
    Spacer(1, 18),
    P("<i>This is a fully synthetic protocol describing a fictional study. No human "
      "subjects were enrolled and no real data are reported. It exists solely to provide "
      "realistic free text and tabular structure from which AI tools can extract metadata "
      "during the ISMB/ECCB CollaborationFest.</i>"),
    PageBreak(),
]

# Synopsis table
f += [P("Protocol Synopsis", "H2b")]
syn = [
    ["Title", STUDY_TITLE],
    ["Design", "Two-arm, parallel-group, randomized pilot trial"],
    ["Population", "Adults 40–65 with prediabetes"],
    ["Intervention", "10-hour time-restricted eating (TRE) window"],
    ["Comparator", "Habitual (unrestricted) eating schedule"],
    ["Primary endpoint", "Change in fasting plasma glucose, baseline to week 12"],
    ["Secondary endpoints", "Change in HbA1c, weight, blood pressure; adherence; safety"],
    ["Sample size", "40 participants (1:1 allocation)"],
    ["Duration", "12 weeks per participant"],
    ["Data standard", "CEDAR templates; registered in the Canopy data hub"],
]
t = Table(syn, colWidths=[1.6*inch, 4.4*inch])
t.setStyle(TableStyle([
    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
    ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#eef5f2")),
    ("VALIGN", (0,0), (-1,-1), "TOP"),
    ("FONTSIZE", (0,0), (-1,-1), 9),
    ("LEFTPADDING", (0,0), (-1,-1), 6), ("RIGHTPADDING", (0,0), (-1,-1), 6),
    ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
]))
f += [t, Spacer(1, 8)]

f += [P("Abbreviations", "H2b"),
      P("TRE — time-restricted eating; FPG — fasting plasma glucose; HbA1c — glycated "
        "hemoglobin; BMI — body mass index; BP — blood pressure; AE — adverse event; "
        "SOP — standard operating procedure; CDE — common data element; FAIR — Findable, "
        "Accessible, Interoperable, Reusable.")]

f += [P("1. Background and Rationale", "H2b"),
      P("Prediabetes — an intermediate state of hyperglycemia in which blood glucose is "
        "elevated but below the diagnostic threshold for type 2 diabetes — affects a large "
        "fraction of middle-aged adults worldwide and is a leading risk factor for progression "
        "to overt diabetes and cardiovascular disease. Lifestyle interventions that improve "
        "glycemic control in this window are therefore of substantial public-health interest."),
      P("Time-restricted eating (TRE) confines daily energy intake to a fixed window, most "
        "commonly 8–10 hours, without explicit calorie counting. Proponents hypothesize that "
        "aligning food intake with circadian rhythms improves insulin sensitivity, reduces "
        "post-prandial glucose excursions, and supports modest weight loss. Early human studies "
        "have reported improvements in fasting glucose and HbA1c, though sample sizes are small "
        "and results are mixed."),
      P("This synthetic pilot models a pragmatic, two-arm randomized comparison of a 10-hour "
        "TRE window against a habitual eating schedule over 12 weeks in adults with prediabetes. "
        "The aim of the (fictional) study is to estimate effect sizes and assess feasibility to "
        "inform a larger confirmatory trial. The aim of <i>this document</i>, within the "
        "CollaborationFest, is to serve as a realistic source of unstructured text from which an "
        "AI-assisted workflow can infer structured, standards-compliant metadata.")]

f += [P("2. Objectives and Endpoints", "H2b"),
      P("<b>Primary objective.</b> To estimate the effect of a 10-hour TRE window, relative to "
        "habitual eating, on the change in fasting plasma glucose from baseline to week 12."),
      P("<b>Secondary objectives.</b> To estimate between-arm differences in change in HbA1c, "
        "body weight and BMI, and systolic blood pressure; to characterize protocol adherence; "
        "and to describe the safety and tolerability of the intervention."),
      P("<b>Endpoints.</b>"),
      bullets([
        "Primary endpoint: change in FPG (mg/dL) from baseline to week 12.",
        "Secondary endpoints: change in HbA1c (%), body weight (kg) and BMI (kg/m^2), and systolic BP (mmHg).",
        "Adherence endpoint: percentage of days the assigned eating window was followed.",
        "Safety endpoint: incidence of adverse events through week 12.",
      ])]

f += [P("3. Study Design", "H2b"),
      P("This is a two-arm, parallel-group, randomized pilot trial conducted across four sites "
        "in the United States, Germany, Spain, and Canada. Following a screening and baseline "
        "visit, eligible participants are randomized 1:1 to the TRE arm or the control arm and "
        "followed for 12 weeks, with a single end-of-study visit at week 12. Randomization uses "
        "a computer-generated sequence stratified by site and sex. Given the behavioral nature "
        "of the intervention, participants and site staff are not blinded; laboratory personnel "
        "processing samples are blinded to arm assignment."),
      P("The total target enrollment is 40 participants. As a pilot, the study is not powered for "
        "confirmatory hypothesis testing; the sample size is chosen to yield stable estimates of "
        "variability and feasibility metrics (recruitment rate, retention, adherence).")]

f += [P("4. Study Population", "H2b"),
      P("The study enrolls community-dwelling adults with prediabetes. Eligibility is assessed at "
        "a screening visit using point-of-care and central laboratory measurements."),
      P("<b>Inclusion criteria.</b>"),
      bullets([
        "Age 40 to 65 years, inclusive, at the time of consent.",
        "BMI between 26.0 and 35.0 kg/m^2.",
        "Documented prediabetes: fasting plasma glucose 100–125 mg/dL or HbA1c 5.7–6.4%.",
        "Willing and able to provide written informed consent and to maintain a daily eating diary.",
      ]),
      P("<b>Exclusion criteria.</b>"),
      bullets([
        "Diagnosed type 1 or type 2 diabetes, or current use of glucose-lowering medication.",
        "Pregnancy, breastfeeding, or plans to become pregnant during the study.",
        "Current or recent (within 3 months) eating disorder.",
        "Shift work or travel that prevents maintaining a fixed daily eating window.",
        "Any condition that, in the investigator's judgment, would compromise safety or data quality.",
      ])]

f += [P("5. Interventions", "H2b"),
      P("<b>TRE arm.</b> Participants confine all caloric intake to a self-selected 10-hour "
        "window, consistently anchored each day (for example 08:00–18:00). Water and non-caloric "
        "beverages are permitted outside the window. Participants are not asked to change food "
        "choices or total intake; only timing is constrained. Adherence is recorded daily in a "
        "diary as whether the window was followed."),
      P("<b>Control arm.</b> Participants maintain their habitual eating schedule and complete "
        "the same diary and assessments. No timing constraint is imposed."),
      P("Both arms receive identical educational materials on the study procedures and identical "
        "assessment schedules, so that any between-arm difference is attributable to eating-window "
        "timing rather than to differing contact or attention.")]

f += [P("6. Study Procedures and Schedule of Assessments", "H2b"),
      P("Participants attend a screening/baseline visit and a week-12 end-of-study visit. The "
        "schedule of assessments is summarized below.")]
sched = [
    ["Assessment", "Screening / Baseline", "Week 12"],
    ["Informed consent", "X", ""],
    ["Eligibility review", "X", ""],
    ["Demographics (age, sex, country)", "X", ""],
    ["Anthropometry (weight, BMI)", "X", "X"],
    ["Fasting plasma glucose", "X", "X"],
    ["HbA1c", "X", "X"],
    ["Blood pressure", "X", "X"],
    ["Randomization", "X", ""],
    ["Adherence diary review", "", "X"],
    ["Adverse-event assessment", "X", "X"],
]
t2 = Table(sched, colWidths=[3.0*inch, 1.7*inch, 1.3*inch])
t2.setStyle(TableStyle([
    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F6F54")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTSIZE", (0,0), (-1,-1), 9),
    ("ALIGN", (1,0), (-1,-1), "CENTER"),
    ("LEFTPADDING", (0,0), (-1,-1), 6), ("TOPPADDING", (0,0), (-1,-1), 3),
    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
]))
f += [t2, Spacer(1, 8),
      P("Blood samples for fasting glucose and HbA1c are collected and handled according to "
        "SOP " + STUDY_ID + "-SOP-001 (provided separately). All measurements are recorded "
        "against the participant identifier in the study database on the day of collection.")]

f += [P("7. Measurements and Data Collection", "H2b"),
      P("The analysis dataset is organized one row per participant, with columns capturing "
        "demographics, the assigned arm, baseline and week-12 laboratory values, adherence, and "
        "safety. Each column has a defined type and, where applicable, a controlled vocabulary or "
        "external identifier authority, documented in the accompanying data dictionary. Variables "
        "of note include:"),
      bullets([
        "Demographics: age (years), sex, and country of the enrolling site.",
        "Condition and arm: the condition under study and the assigned study arm.",
        "Laboratory: fasting plasma glucose at baseline and week 12, HbA1c, and systolic blood pressure.",
        "Anthropometry: BMI at baseline.",
        "Behavioral: protocol adherence, expressed as a percentage of adherent days.",
        "Safety: whether any adverse event was reported, and whether the participant completed the study.",
        "Provenance: the enrolling investigator's ORCID iD, the reference publication's PubMed ID, and the protocol DOI.",
      ]),
      P("Dates (enrollment and final visit) are recorded in ISO-8601 format. Boolean fields "
        "(medication use, adverse event, study completion) are recorded as true/false. These "
        "varied field types are intentional, so that downstream metadata description must handle "
        "numeric, date, boolean, controlled-term, and identifier fields rather than treating "
        "everything as free text.")]

f += [P("8. Statistical Analysis", "H2b"),
      P("The primary endpoint, change in fasting plasma glucose from baseline to week 12, is "
        "compared between arms using a two-sample t-test. As a sensitivity analysis, a linear "
        "model adjusts for the baseline value, age, and sex. Secondary continuous endpoints are "
        "analyzed analogously. Adherence is summarized descriptively, and the relationship between "
        "adherence and glycemic change is explored. Safety is summarized by counts and proportions "
        "of participants reporting adverse events by arm."),
      P("Because this is a pilot, all inferential results are considered exploratory and are "
        "reported with confidence intervals rather than emphasizing p-value thresholds. Missing "
        "week-12 values (for participants who did not complete the study) are handled by a complete-"
        "case primary analysis, with a sensitivity analysis using multiple imputation.")]

f += [P("9. Data Management, Metadata, and FAIR Sharing", "H2b"),
      P("Study-level and dataset-level metadata are described using CEDAR templates. The study is "
        "registered in the Canopy data hub, where the dataset, data dictionary, and these "
        "documents are bundled, validated, and — upon curator approval — published according to "
        "the study's access level. Controlled values (sex, country, condition, study arm) are "
        "bound to ontology terms via BioPortal so that the dataset is interoperable and machine-"
        "queryable, and external identifiers (ORCID, PubMed ID, DOI) anchor provenance to "
        "authoritative registries."),
      P("As a synthetic dataset released for tooling and education, the data carry an open license "
        "and may be freely reused.")]

f += [P("10. Ethics", "H2b"),
      P("Because all subjects and results are fabricated, no human-subjects research was conducted "
        "and no ethics approval applies. In a real study of this design, the protocol would be "
        "reviewed by an institutional review board at each site, all participants would provide "
        "written informed consent, and the trial would be registered in a public registry before "
        "enrollment.")]

f += [P("11. Adverse Events: Definitions and Reporting", "H2b"),
      P("An adverse event (AE) is any untoward medical occurrence in a participant during the study, "
        "whether or not it is considered related to the intervention. Because TRE is a behavioral "
        "intervention, anticipated events are generally mild and may include transient hunger, "
        "headache, irritability, or light-headedness, particularly in the first two weeks as "
        "participants adjust to the eating window."),
      P("Site staff assess for AEs at each contact and review the participant diary at the week-12 "
        "visit. Each event is recorded with its onset date, severity (mild, moderate, severe), "
        "the investigator's assessment of relatedness to the intervention, the action taken, and "
        "the outcome. In the analysis dataset, the presence of any reported AE for a participant is "
        "captured by the boolean field adverse_event; the full event-level detail would be held in a "
        "separate AE log in a complete study."),
      P("<b>Serious adverse events.</b> A serious adverse event (SAE) is one that results in death, "
        "is life-threatening, requires hospitalization, or results in persistent disability. Any SAE "
        "would be reported to the site IRB and the sponsor within the timelines specified by local "
        "regulation. No SAEs are expected for an intervention of this nature, and — as a synthetic "
        "study — none are reported here.")]

f += [P("12. Informed Consent Process", "H2b"),
      P("In a real conduct of this study, written informed consent would be obtained from every "
        "participant before any study procedure. The consent discussion would cover the purpose, "
        "the random assignment, the eating-window requirement, the blood draws, the expected "
        "duration, the voluntary nature of participation, the right to withdraw at any time without "
        "penalty, and how personal data would be stored, shared, and protected. Participants would "
        "receive a copy of the signed form and have the opportunity to ask questions before and "
        "during the study."),
      P("Because this protocol describes a synthetic study, no consent was obtained and no personal "
        "data exist; all participant records are computer-generated.")]

f += [P("13. Recruitment, Retention, and Withdrawal", "H2b"),
      P("Participants would be recruited from primary-care referrals and community advertising at "
        "each of the four sites, with screening to confirm prediabetes and eligibility. To support "
        "retention over the 12-week period, sites provide reminders, a simple diary, and a single "
        "follow-up visit rather than a burdensome schedule."),
      P("A participant may withdraw at any time, and the investigator may withdraw a participant for "
        "safety or eligibility reasons. For participants who do not complete the study, baseline data "
        "are retained and week-12 values are recorded as missing. In the analysis dataset, study "
        "completion is captured by the boolean field completed_study, and missing week-12 laboratory "
        "values are represented as empty cells.")]

f += [P("14. Study Organization and Governance", "H2b"),
      P("The study is coordinated by a lead investigator with a site principal investigator at each "
        "of the four participating sites. A small operations team maintains the protocol, the SOPs, "
        "the randomization sequence, and the study database. Each enrolling investigator is identified "
        "by an ORCID iD, which is carried into the dataset (investigator_orcid) so that records can be "
        "linked unambiguously to the responsible investigator."),
      P("Roles relevant to data submission map onto the Canopy platform: a Data Submitter registers "
        "the study and uploads the dataset and documents, and a Data Curator reviews the submission "
        "before publication. This separation of duties preserves an independent review step.")]

f += [P("15. Data Monitoring and Quality Control", "H2b"),
      P("Data quality is maintained through source-document verification, range and consistency checks "
        "on key variables (for example, week-12 glucose values are checked against plausible "
        "physiological bounds), and reconciliation of the adherence diary against the recorded "
        "adherence percentage. Laboratory measurements follow the sample-handling SOP and include "
        "two levels of quality-control material per analytical batch."),
      P("Validation continues at submission time: when the dataset, its data dictionary, and these "
        "documents are uploaded to Canopy, the platform validates each bundle against the required "
        "metadata template and data-dictionary specification before a curator reviews and approves it.")]

f += [P("16. Limitations", "H2b"),
      P("This pilot has several limitations by design. The sample size is small and the study is not "
        "powered for confirmatory inference, so estimates of effect should be interpreted with wide "
        "uncertainty. The behavioral intervention cannot be blinded, which may introduce reporting "
        "and expectation effects; adherence is self-reported through a daily diary and may be subject "
        "to recall and social-desirability bias. The 12-week duration captures short-term glycemic "
        "change but cannot speak to durability or to progression to diabetes. Finally, recruitment "
        "from four sites in four countries supports generalizability across settings but introduces "
        "between-site heterogeneity in measurement and population that a pilot can describe but not "
        "fully adjust for."),
      P("For the purposes of the CollaborationFest, these limitations are immaterial to the data's "
        "function: the value of this synthetic study is the realistic structure and free text it "
        "provides for metadata extraction, not the validity of its (fabricated) clinical findings.")]

f += [P("17. Study Timeline and Milestones", "H2b"),
      P("The study is conducted over a single calendar period, with rolling enrollment followed by a "
        "12-week per-participant follow-up. Indicative milestones are summarized below.")]
tl = [
    ["Milestone", "Indicative timing"],
    ["Protocol finalized and sites activated", "Month 0"],
    ["First participant enrolled", "Month 1"],
    ["Enrollment complete (40 participants)", "Month 2"],
    ["Last participant week-12 visit", "Month 5"],
    ["Database lock and dataset assembled", "Month 6"],
    ["Metadata described and submitted to Canopy", "Month 6"],
]
ttl = Table(tl, colWidths=[3.6*inch, 2.4*inch])
ttl.setStyle(TableStyle([
    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F6F54")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTSIZE", (0,0), (-1,-1), 9),
    ("LEFTPADDING", (0,0), (-1,-1), 6), ("TOPPADDING", (0,0), (-1,-1), 3),
    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
]))
f += [ttl, Spacer(1, 8),
      P("At database lock, the analysis dataset, its data dictionary, this protocol, and the "
        "sample-handling SOP are assembled into a submission package. The metadata-description step "
        "— filling the Canopy Study template and a domain-specific template, and registering the "
        "study — is the activity this CollaborationFest project seeks to assist with AI tooling.")]

f += [PageBreak(),
      P("Appendix A. Data Dictionary", "H2b"),
      P("The analysis dataset contains 20 columns, one row per participant. The accompanying "
        "spreadsheet includes this dictionary as a separate sheet; it is reproduced here so the "
        "structure is visible from the protocol alone.")]
appx = [["Column", "Type", "Controlled by / authority"]]
for name, ftype, auth, desc in SCHEMA:
    appx.append([name, ftype, auth or "—"])
ta = Table(appx, colWidths=[2.3*inch, 1.2*inch, 2.5*inch], repeatRows=1)
ta.setStyle(TableStyle([
    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#cccccc")),
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F6F54")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTSIZE", (0,0), (-1,-1), 9),
    ("LEFTPADDING", (0,0), (-1,-1), 6), ("TOPPADDING", (0,0), (-1,-1), 3),
    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
]))
f += [ta, Spacer(1, 6),
      P("Field descriptions, units, and ontology references for each column are given in the "
        "data_dictionary sheet of SPbE-2026_dataset.xlsx.")]

f += [P("Appendix B. Glossary and Controlled-Term Notes", "H2b"),
      P("The following definitions clarify terms used in this protocol and indicate the controlled "
        "vocabularies to which the corresponding dataset fields are bound. These notes are intended "
        "to support automated metadata description: each controlled field should resolve to a term "
        "in the named ontology rather than remain free text."),
      bullets([
        "<b>Prediabetes</b> — an intermediate hyperglycemic state (fasting glucose 100–125 mg/dL or "
        "HbA1c 5.7–6.4%). Dataset field <i>condition</i> binds to MONDO:0005827.",
        "<b>Time-restricted eating (TRE)</b> — confining caloric intake to a fixed daily window. "
        "Captured by the <i>study_arm</i> field (TRE vs. control), bound to an NCIT study-arm term.",
        "<b>Fasting plasma glucose</b> — plasma glucose after ≥8 hours fasting, in mg/dL. Fields "
        "<i>baseline_glucose_mg_dl</i> and <i>week12_glucose_mg_dl</i>, bound to NCIT:C105585.",
        "<b>HbA1c</b> — glycated hemoglobin, reflecting average glycemia over ~3 months, in percent. "
        "Field <i>hba1c_pct</i>, bound to NCIT:C64849.",
        "<b>Country</b> — country of the enrolling site, bound to a geographic ontology branch "
        "(GAZ) / ISO-3166-1, rather than entered as arbitrary text.",
        "<b>ORCID iD</b> — a persistent identifier for a researcher (https://orcid.org). Field "
        "<i>investigator_orcid</i> anchors each record to the responsible investigator.",
        "<b>PubMed ID (PMID)</b> — a persistent identifier for a publication in PubMed. Field "
        "<i>reference_pmid</i> links to the reference publication.",
        "<b>DOI</b> — a digital object identifier for a resource (https://doi.org). Field "
        "<i>protocol_doi</i> identifies this protocol.",
      ]),
      P("Binding these fields to recognized authorities is what makes the resulting dataset "
        "Findable, Accessible, Interoperable, and Reusable (FAIR): a value such as a country or a "
        "condition becomes globally unambiguous, and provenance can be followed to an authoritative "
        "registry.")]

f += [P("References", "H2b"),
      P("Synthetic protocol; references are illustrative. A representative review of time-restricted "
        "eating and metabolic health is indexed in PubMed under PMID " + REFERENCE_PMID + ". This "
        "protocol is identified by DOI " + PROTOCOL_DOI + ".")]

build_pdf(OUT / "SPbE-2026_protocol.pdf", f)

# ---------------------------------------------------------------- SOP PDF
f = [
    P("Standard Operating Procedure", "H"),
    P("<b>SOP: Fasting Plasma Glucose and HbA1c Sample Collection and Handling</b>"),
    P(f"SOP ID: {STUDY_ID}-SOP-001 &nbsp;|&nbsp; Version 1.0 &nbsp;|&nbsp; Synthetic example."),
    Spacer(1, 10),
    P("1. Purpose", "H2b"),
    P("To standardize the collection, labeling, processing, and storage of venous blood samples used to "
      "measure fasting plasma glucose and HbA1c, ensuring comparable results across visits and subjects."),
    P("2. Scope", "H2b"),
    P(f"Applies to all study staff collecting biospecimens for study {STUDY_ID} at baseline and week 12."),
    P("3. Materials", "H2b"),
    bullets([
        "Sodium fluoride / potassium oxalate tubes (grey-top) for glucose.",
        "EDTA tubes (lavender-top) for HbA1c.",
        "Pre-printed barcode labels linked to subject_id.",
        "Calibrated centrifuge, refrigerated transport box, validated -80 C freezer.",
    ]),
    P("4. Procedure", "H2b"),
    ListFlowable([ListItem(P(t)) for t in [
        "Confirm the participant has fasted >= 8 hours; record last food intake time.",
        "Collect blood by standard venipuncture into the grey-top tube first, then the lavender-top tube.",
        "Invert each tube gently 8-10 times immediately after collection. Do not shake.",
        "Label each tube with the barcode and record visit_date and time.",
        "Centrifuge the glucose tube within 30 minutes (1500 x g, 10 min, 4 C); separate plasma.",
        "Store HbA1c whole blood at 4 C and analyze within 72 hours; archive plasma aliquots at -80 C.",
        "Log every sample in the LIMS against the subject_id before end of day.",
    ]], bulletType="1", leftIndent=18),
    P("5. Quality Control", "H2b"),
    P("Run two levels of commercial control material with each analytical batch. Reject and recollect any "
      "sample with visible hemolysis or an unbroken cold chain. Document deviations on the deviation log."),
    P("6. Records", "H2b"),
    P("Retain collection logs, instrument calibration records, and QC results. These records map directly "
      "to the dataset fields visit_date, baseline_glucose_mg_dl, week12_glucose_mg_dl, and hba1c_pct."),
]
build_pdf(OUT / "SPbE-2026_SOP_sample-collection.pdf", f)
print("DONE")
